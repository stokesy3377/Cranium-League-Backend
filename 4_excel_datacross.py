import openpyxl

# Load workbook with data_only=True to read actual values (not formulas)
workbook = openpyxl.load_workbook('CL_S1.xlsx', data_only=True)
worksheet = workbook.active

# Separate workbook for writing changes (without data_only)
writebook = openpyxl.load_workbook('CL_S1.xlsx')
writesheet = writebook.active

def get_numeric(ws, cell):
    val = ws[cell].value
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        val = val.replace(",", "").strip()
        try:
            return float(val)
        except ValueError:
            return 0
    return 0

x = 5
while x <= 13:
    current_balance = get_numeric(worksheet, 'I' + str(x))
    previous_balance = get_numeric(worksheet, 'H' + str(x))

    writesheet['H' + str(x)].value = current_balance
    writesheet['G' + str(x)].value = 0
    writesheet['N' + str(x)].value = 0

    x += 1

writebook.save('CL_S1.xlsx')
print("✅ Data updated successfully using formula-evaluated values.")