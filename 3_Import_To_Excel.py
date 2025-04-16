import openpyxl 
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl import workbook

# Ask for the date input
input_date = input("Enter the Match Number (prefix 'M'): ")

# Load the workbook and the active worksheet
workbook = openpyxl.load_workbook('CL_S1.xlsx')
worksheet = workbook.active

# Iterate through rows and columns to find the date
for row in worksheet.iter_rows():
    for cell in row:
        if cell.value == input_date:
            # Return the cell coordinates
            column_letter = get_column_letter(cell.column)
            row_number = cell.row
            print(f"The date {input_date} was found at {column_letter}{row_number}")
            # Optional: Return the cell reference as a string
            cell_reference = coordinate_from_string(f"{column_letter}{row_number}")
            print(f"Cell reference: {cell_reference}")
            break
    else:
        continue
    break
else:
    print("The date was not found in the worksheet.")

# Save the workbook (optional)
workbook.save('CL_S1.xlsx')

colN = column_letter
rowN = row_number

colN = colN.upper()


from CL_data_mod import lst1
print(colN)

x=1
for bal in lst1:
    worksheet[colN+str(rowN+x)] = bal
    x+=1

workbook.save('CL_S1.xlsx')

